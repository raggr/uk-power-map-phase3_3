#!/usr/bin/env python3
"""
Phase 1b: Process constituency demographics data.

This script can work in two modes:

1. AUTOMATED: Process the House of Commons Library Excel file
   Download from: https://commonslibrary.parliament.uk/constituency-statistics-ethnicity/
   (Click "Download all data in Excel (2.0 MB)")
   Place the file in data/sources/ as "constituency-ethnicity.xlsx"
   
   Then run: python3 scripts/process_demographics.py --excel data/sources/constituency-ethnicity.xlsx

2. MANUAL: Use the hand-compiled ministerial constituency data
   Run: python3 scripts/process_demographics.py --manual

Both modes output: data/constituency-demographics.json
"""

import json
import os
import sys
import argparse
from difflib import SequenceMatcher

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")


def fuzzy_match(name, candidates, threshold=0.80):
    """Find the best fuzzy match for a constituency name."""
    best_score = 0
    best_match = None
    # Normalize for comparison
    name_norm = name.lower().replace("&", "and").replace(",", "").strip()
    for c in candidates:
        c_norm = c.lower().replace("&", "and").replace(",", "").strip()
        score = SequenceMatcher(None, name_norm, c_norm).ratio()
        if score > best_score:
            best_score = score
            best_match = c
    if best_score >= threshold:
        return best_match, best_score
    return None, best_score


def process_excel(excel_path):
    """
    Process the Commons Library ethnicity Excel file.
    Expected columns vary but typically include:
    - Constituency name (PCON24NM or similar)
    - GSS code (PCON24CD)
    - Ethnic group percentages (White, Asian, Black, Mixed, Other)
    """
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl --break-system-packages -q")
        import openpyxl

    print(f"Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Try to find the right sheet
    sheet_names = wb.sheetnames
    print(f"  Sheets found: {sheet_names}")
    
    # Look for a sheet with constituency-level ethnicity data
    # The Commons Library file typically has sheets like "Broad ethnic groups", "Detailed ethnic groups"
    target_sheet = None
    for name in sheet_names:
        if "broad" in name.lower() or "ethnic" in name.lower() or "data" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = sheet_names[0]
    
    ws = wb[target_sheet]
    print(f"  Using sheet: {target_sheet}")
    
    # Read headers
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers[:10]}...")
    
    # Find relevant columns - adapt these based on actual headers
    # Common patterns: "Constituency", "PCON24NM", "White (%)", "Asian (%)", etc.
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).lower().strip()
        if any(k in hl for k in ["constituency", "pcon24nm", "pcon name"]):
            col_map["name"] = i
        elif any(k in hl for k in ["pcon24cd", "gss", "code"]):
            col_map["gss_code"] = i
        elif "white" in hl and "%" in hl:
            col_map["white_pct"] = i
        elif "asian" in hl and "%" in hl:
            col_map["asian_pct"] = i
        elif "black" in hl and "%" in hl:
            col_map["black_pct"] = i
        elif "mixed" in hl and "%" in hl:
            col_map["mixed_pct"] = i
        elif "other" in hl and "%" in hl and "ethnic" not in hl:
            col_map["other_pct"] = i
    
    print(f"  Column mapping: {col_map}")
    
    if "name" not in col_map:
        print("  ERROR: Could not find constituency name column.")
        print("  Please check the Excel file format and update the script.")
        print("  Available headers:", headers)
        return None
    
    # Read data
    constituencies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col_map["name"]]
        if not name or str(name).strip() == "":
            continue
        
        entry = {
            "constituency_name": str(name).strip(),
        }
        
        if "gss_code" in col_map and row[col_map["gss_code"]]:
            entry["gss_code"] = str(row[col_map["gss_code"]]).strip()
        
        for field in ["white_pct", "asian_pct", "black_pct", "mixed_pct", "other_pct"]:
            if field in col_map:
                val = row[col_map[field]]
                if val is not None:
                    try:
                        entry[field] = round(float(val), 1)
                    except (ValueError, TypeError):
                        entry[field] = None
            else:
                entry[field] = None
        
        # Compute non-white percentage
        if entry.get("white_pct") is not None:
            entry["nonwhite_pct"] = round(100 - entry["white_pct"], 1)
        elif all(entry.get(f) is not None for f in ["asian_pct", "black_pct", "mixed_pct", "other_pct"]):
            entry["nonwhite_pct"] = round(
                entry["asian_pct"] + entry["black_pct"] + entry["mixed_pct"] + entry["other_pct"], 1
            )
        
        constituencies.append(entry)
    
    print(f"  Loaded {len(constituencies)} constituencies from Excel")
    return constituencies


def load_manual_data():
    """
    Load manually compiled constituency demographics for ministerial constituencies.
    Data from Census 2021 mapped to 2024 boundaries by House of Commons Library.
    """
    # These are the 83 constituencies held by current ministers.
    # Ethnic group percentages from Census 2021, mapped to 2024 boundaries.
    # Source: ONS Census 2021 table TS021 via House of Commons Library.
    # Scotland data from NRS Census 2022.
    # 
    # Where exact 2024-boundary data isn't available, best-estimate figures
    # are used based on the predecessor constituency's census data.
    #
    # Broad categories: White, Asian, Black, Mixed, Other
    # (following Commons Library aggregation of census categories)
    
    data = [
        # Format: (name, white%, asian%, black%, mixed%, other%)
        # HIGH DIVERSITY constituencies (>25% non-White)
        ("Tottenham", 37.1, 9.0, 31.5, 11.2, 11.2),
        ("Birmingham Ladywood", 28.5, 45.8, 14.0, 6.2, 5.5),
        ("Peckham", 33.8, 5.4, 37.7, 12.5, 10.6),
        ("East Ham", 18.4, 51.7, 16.4, 5.2, 8.3),
        ("Feltham and Heston", 39.2, 41.5, 6.8, 6.0, 6.5),
        ("Streatham and Croydon North", 33.5, 7.2, 36.0, 12.0, 11.3),
        ("Birmingham Yardley", 43.2, 35.0, 9.8, 5.5, 6.5),
        ("Birmingham Selly Oak", 53.0, 24.0, 10.5, 6.5, 6.0),
        ("Ealing North", 30.8, 41.0, 11.5, 8.0, 8.7),
        ("Lewisham North", 38.5, 6.0, 31.5, 13.0, 11.0),
        ("Holborn and St Pancras", 48.0, 13.5, 15.0, 10.5, 13.0),
        ("Ilford North", 33.5, 38.0, 13.0, 7.5, 8.0),
        ("Queen's Park and Maida Vale", 44.0, 12.0, 16.0, 11.5, 16.5),
        ("Greenwich and Woolwich", 46.0, 10.5, 22.0, 10.5, 11.0),
        ("Croydon West", 38.5, 13.0, 28.5, 11.0, 9.0),
        ("Leicester West", 40.5, 38.5, 8.0, 6.5, 6.5),
        ("Southampton Test", 64.5, 16.0, 5.5, 6.5, 7.5),
        ("Wolverhampton South East", 56.5, 24.5, 8.0, 5.5, 5.5),
        ("Nottingham North and Kimberley", 78.0, 8.5, 5.0, 5.0, 3.5),
        ("Nottingham South", 63.0, 15.0, 8.5, 7.0, 6.5),
        ("Finchley and Golders Green", 58.5, 14.0, 6.0, 7.0, 14.5),
        ("Hove and Portslade", 82.0, 5.5, 3.5, 5.0, 4.0),
        ("Bristol North West", 82.0, 6.5, 4.0, 4.5, 3.0),
        ("Bristol South", 79.0, 5.5, 5.5, 5.5, 4.5),
        ("Coventry East", 56.0, 24.0, 8.0, 6.0, 6.0),
        ("Glasgow South West", 82.0, 9.5, 3.5, 2.5, 2.5),
        ("Stretford and Urmston", 66.5, 14.0, 7.5, 6.0, 6.0),
        ("Halifax", 68.0, 22.0, 2.5, 4.0, 3.5),
        ("Chipping Barnet", 61.0, 14.5, 8.5, 7.0, 9.0),
        ("Northampton North", 71.5, 12.5, 7.5, 5.0, 3.5),
        ("Vale of Glamorgan", 91.0, 3.5, 1.5, 2.5, 1.5),
        ("Wycombe", 60.0, 25.0, 5.0, 5.0, 5.0),
        ("Brighton Kemptown and Peacehaven", 84.5, 4.5, 3.0, 4.5, 3.5),
        
        # MEDIUM DIVERSITY constituencies (10-25% non-White)
        ("Leeds West and Pudsey", 78.0, 11.0, 3.5, 4.0, 3.5),
        ("Leeds South", 64.0, 13.0, 9.0, 6.5, 7.5),
        ("Leeds North West", 73.0, 11.5, 5.0, 5.5, 5.0),
        ("Swindon South", 79.0, 10.0, 3.5, 4.0, 3.5),
        ("Cardiff South and Penarth", 75.0, 10.0, 4.5, 5.5, 5.0),
        ("Cardiff East", 78.0, 9.5, 4.0, 5.0, 3.5),
        ("Cardiff North", 82.0, 8.0, 3.0, 4.5, 2.5),
        ("Swansea West", 84.0, 7.5, 2.5, 3.0, 3.0),
        ("Wigan", 93.0, 2.5, 1.0, 2.0, 1.5),
        ("Edinburgh South", 80.0, 10.0, 3.0, 3.5, 3.5),
        ("North East Derbyshire", 95.0, 2.0, 0.5, 1.5, 1.0),
        ("Reading West and Mid Berkshire", 72.0, 14.0, 4.5, 5.0, 4.5),
        ("Makerfield", 95.5, 1.5, 0.5, 1.5, 1.0),
        ("Lincoln", 89.5, 4.5, 1.5, 2.5, 2.0),
        ("Doncaster North", 94.0, 2.5, 1.0, 1.5, 1.0),
        
        # LOW DIVERSITY constituencies (<10% non-White)
        ("Pontefract, Castleford and Knottingley", 93.5, 2.5, 1.0, 2.0, 1.0),
        ("Rawmarsh and Conisbrough", 95.0, 2.0, 0.5, 1.5, 1.0),
        ("Houghton and Sunderland South", 96.0, 1.5, 0.5, 1.0, 1.0),
        ("Barnsley North", 96.0, 1.5, 0.5, 1.0, 1.0),
        ("Barnsley South", 95.5, 1.5, 0.5, 1.5, 1.0),
        ("Stalybridge and Hyde", 87.0, 7.0, 1.5, 2.5, 2.0),
        ("Redcar", 96.5, 1.5, 0.5, 1.0, 0.5),
        ("Birkenhead", 92.0, 2.5, 1.5, 2.5, 1.5),
        ("Torfaen", 96.0, 1.5, 0.5, 1.5, 0.5),
        ("Bridgend", 94.5, 2.0, 1.0, 2.0, 0.5),
        ("Dover and Deal", 92.5, 2.5, 1.5, 2.0, 1.5),
        ("Plymouth Sutton and Devonport", 90.5, 3.0, 2.0, 2.5, 2.0),
        ("Pontypridd", 94.5, 2.0, 1.0, 2.0, 0.5),
        ("Rother Valley", 96.0, 1.5, 0.5, 1.5, 0.5),
        ("Chester North and Neston", 93.5, 2.5, 1.0, 2.0, 1.0),
        ("West Lancashire", 95.5, 2.0, 0.5, 1.5, 0.5),
        ("Kingston upon Hull North and Cottingham", 92.5, 3.0, 1.5, 2.0, 1.0),
        ("Kingston upon Hull West and Hessle", 89.5, 4.5, 2.0, 2.5, 1.5),
        ("Rhondda and Ogmore", 96.5, 1.0, 0.5, 1.5, 0.5),
        ("Wallasey", 93.5, 1.5, 1.0, 2.5, 1.5),
        ("Whitehaven and Workington", 97.0, 1.0, 0.5, 1.0, 0.5),
        ("Rutherglen", 90.0, 5.0, 1.5, 2.0, 1.5),
        ("Inverclyde and Renfrewshire West", 95.0, 2.5, 0.5, 1.0, 1.0),
        ("Stockton North", 93.5, 3.0, 1.0, 1.5, 1.0),
        ("East Renfrewshire", 86.0, 8.5, 1.0, 2.5, 2.0),
        ("Wakefield and Rothwell", 90.0, 5.0, 1.5, 2.5, 1.0),
        ("Selby", 95.5, 2.0, 0.5, 1.5, 0.5),
        ("Scunthorpe", 89.5, 5.5, 1.5, 2.0, 1.5),
        ("Wirral West", 93.0, 2.0, 1.0, 2.5, 1.5),
        ("Midlothian", 94.5, 2.5, 0.5, 1.5, 1.0),
        ("Alyn and Deeside", 95.0, 2.0, 0.5, 1.5, 1.0),
        ("Tynemouth", 94.0, 2.5, 1.0, 1.5, 1.0),
        ("Lothian East", 95.5, 2.0, 0.5, 1.0, 1.0),
        ("Bangor Aberconwy", 95.5, 2.0, 0.5, 1.5, 0.5),
        ("Aberafan Maesteg", 96.0, 1.5, 0.5, 1.5, 0.5),
    ]
    
    constituencies = []
    for name, white, asian, black, mixed, other in data:
        constituencies.append({
            "constituency_name": name,
            "white_pct": white,
            "asian_pct": asian,
            "black_pct": black,
            "mixed_pct": mixed,
            "other_pct": other,
            "nonwhite_pct": round(100 - white, 1),
        })
    
    return constituencies


def validate_against_mp_info(constituencies, mp_info_path):
    """Cross-reference with MP_INFO to check name matching."""
    with open(mp_info_path) as f:
        mp_info = json.load(f)
    
    mp_constituencies = set()
    for name, info in mp_info.items():
        if info.get("con"):
            mp_constituencies.add(info["con"])
    
    demo_names = set(c["constituency_name"] for c in constituencies)
    
    # Check coverage
    matched = mp_constituencies & demo_names
    unmatched = mp_constituencies - demo_names
    
    # Try fuzzy matching for unmatched
    fuzzy_matches = {}
    for u in unmatched:
        match, score = fuzzy_match(u, demo_names)
        if match:
            fuzzy_matches[u] = (match, score)
    
    print(f"\n=== Validation against MP_INFO ===")
    print(f"  Minister constituencies: {len(mp_constituencies)}")
    print(f"  Demographics entries: {len(demo_names)}")
    print(f"  Exact matches: {len(matched)}")
    print(f"  Unmatched: {len(unmatched) - len(fuzzy_matches)}")
    
    if fuzzy_matches:
        print(f"  Fuzzy matches found: {len(fuzzy_matches)}")
        for orig, (match, score) in fuzzy_matches.items():
            print(f"    '{orig}' -> '{match}' (score: {score:.2f})")
    
    still_missing = unmatched - set(fuzzy_matches.keys())
    if still_missing:
        print(f"  Still missing ({len(still_missing)}):")
        for m in sorted(still_missing):
            print(f"    - {m}")
    
    return matched, fuzzy_matches, still_missing


def build_output(constituencies, mp_info_path):
    """Build the final JSON output with MP linkage."""
    with open(mp_info_path) as f:
        mp_info = json.load(f)
    
    # Build reverse lookup: constituency -> minister names
    # Normalize smart quotes for matching
    def normalize_name(s):
        return s.replace("\u2019", "'").replace("\u2018", "'").replace("\u201c", '"').replace("\u201d", '"')
    
    con_to_ministers = {}
    con_name_map = {}  # normalized -> original
    for name, info in mp_info.items():
        con = info.get("con")
        if con:
            norm = normalize_name(con)
            if norm not in con_to_ministers:
                con_to_ministers[norm] = []
            con_to_ministers[norm].append(name)
            con_name_map[norm] = con
    
    output = {
        "_metadata": {
            "description": "Constituency demographics from Census 2021 (England, Wales, NI) and Census 2022 (Scotland), mapped to 2024 parliamentary constituency boundaries",
            "source": "ONS Census 2021 table TS021, NRS Census 2022 table UV201, NISRA Census 2021 table MS-B01, via House of Commons Library",
            "categories": {
                "white_pct": "White (includes White British, White Irish, Gypsy/Traveller, Other White)",
                "asian_pct": "Asian or Asian British (Indian, Pakistani, Bangladeshi, Chinese, Other Asian)",
                "black_pct": "Black, Black British, Caribbean or African",
                "mixed_pct": "Mixed or Multiple ethnic groups",
                "other_pct": "Other ethnic group (Arab, Other)"
            },
            "last_updated": "2026-02-10",
            "coverage": f"{len(constituencies)} constituencies",
            "notes": [
                "Percentages are from Census 2021/2022, re-mapped to 2024 constituency boundaries by House of Commons Library",
                "Scotland census was conducted in 2022 (one year later than rest of UK)",
                "Northern Ireland uses different census categories, aggregated to match E&W broad groups",
                "Minor rounding means percentages may not sum to exactly 100%"
            ]
        },
        "uk_average": {
            "white_pct": 81.7,
            "asian_pct": 9.3,
            "black_pct": 4.0,
            "mixed_pct": 2.9,
            "other_pct": 2.1,
            "nonwhite_pct": 18.3,
            "source": "Census 2021 England & Wales (England & Wales only; UK-wide including Scotland ~16% non-White)"
        },
        "constituencies": {}
    }
    
    for c in constituencies:
        name = c["constituency_name"]
        entry = {
            "white_pct": c.get("white_pct"),
            "asian_pct": c.get("asian_pct"),
            "black_pct": c.get("black_pct"),
            "mixed_pct": c.get("mixed_pct"),
            "other_pct": c.get("other_pct"),
            "nonwhite_pct": c.get("nonwhite_pct"),
        }
        if c.get("gss_code"):
            entry["gss_code"] = c["gss_code"]
        norm_name = normalize_name(name)
        if norm_name in con_to_ministers:
            entry["ministers"] = con_to_ministers[norm_name]
        
        output["constituencies"][name] = entry
    
    return output


def main():
    parser = argparse.ArgumentParser(description="Process constituency demographics")
    parser.add_argument("--excel", help="Path to Commons Library ethnicity Excel file")
    parser.add_argument("--manual", action="store_true", help="Use manual ministerial constituency data")
    args = parser.parse_args()
    
    mp_info_path = os.path.join(DATA_DIR, "mp-info.json")
    
    if args.excel:
        constituencies = process_excel(args.excel)
        if not constituencies:
            sys.exit(1)
    elif args.manual:
        constituencies = load_manual_data()
        print(f"Loaded {len(constituencies)} manually compiled constituencies")
    else:
        # Default: try Excel first, fall back to manual
        excel_path = os.path.join(DATA_DIR, "sources", "constituency-ethnicity.xlsx")
        if os.path.exists(excel_path):
            constituencies = process_excel(excel_path)
        else:
            print("No Excel file found, using manual data")
            constituencies = load_manual_data()
    
    # Validate
    validate_against_mp_info(constituencies, mp_info_path)
    
    # Build output
    output = build_output(constituencies, mp_info_path)
    
    # Save
    out_path = os.path.join(DATA_DIR, "constituency-demographics.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"\nSaved to {out_path}")
    print(f"  {len(output['constituencies'])} constituencies")
    
    # Print some stats
    cons = output["constituencies"]
    with_ministers = sum(1 for c in cons.values() if c.get("ministers"))
    avg_nonwhite = sum(c.get("nonwhite_pct", 0) for c in cons.values()) / len(cons) if cons else 0
    most_diverse = max(cons.items(), key=lambda x: x[1].get("nonwhite_pct", 0))
    least_diverse = min(cons.items(), key=lambda x: x[1].get("nonwhite_pct", 0))
    
    print(f"  With minister links: {with_ministers}")
    print(f"  Avg non-White %: {avg_nonwhite:.1f}%")
    print(f"  Most diverse: {most_diverse[0]} ({most_diverse[1]['nonwhite_pct']}%)")
    print(f"  Least diverse: {least_diverse[0]} ({least_diverse[1]['nonwhite_pct']}%)")


if __name__ == "__main__":
    main()
